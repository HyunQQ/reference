# <Title>
# openpyxl을 통해 excel 보고서 만들기

# <Install>
# pip install openpyxl

# <Using Version>
# Python 3.6.5
# pip freeze | findstr psycopg2
# openpyxl==2.5.3

# <Reference>
# doc : https://openpyxl.readthedocs.io/en/stable/


# <Example Select Code>
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.cell import get_column_letter

from datetime import date
import pandas as pd

file_nm = "excel-test.xlsx"

wb = Workbook()             
ws_report = wb.active       #기본 시트 활성화
ws_report.title="Report"    #기본 시트에 Report라는 이름 할당
# ws_report = wb.create_sheet("Report",0)  # Qsheet1 이름의 시트를 1번째 위치에 생성
ws_data = wb.create_sheet("Data",1)  # Qsheet2 이름의 시트를 2번째 위치에 생성
# 위치를 건너뛰고 생성했을 경우 sheet이름의 중산 시트가 1개 생성되고 설정 이름의 시트가 생성된다. 
# 순서가 잘못되면 원하는 위치로 생성되지 않는 것으로 보임.
# ex)ws = wb.create_sheet("Qsheet5",5)  -> 생성시트 (Qsheet1, Qsheet2, sheet,  Qsheet5)


######################### Report sheet #########################
# cell for title 
ws_report.merge_cells('A1:H1')
ws_report['A1'] = 'Report'
title = ws_report['A1']

title.fill = PatternFill("solid", fgColor="DDDDDD")
title.font = Font(name='맑은 고딕', size = 15, bold=True)
title.alignment = Alignment(horizontal='center', vertical='center')

#################################################################


######################### Data sheet #########################
# cell for title 

# cell for dataframe data


data = pd.DataFrame([
    ['Date', 'Batch 1', 'longgggg111data'],
    [date(2015,9, 1), 40, None],
    [date(2015,9, 2), 40, None],
    [date(2015,9, 3), 50, None],
    [date(2015,9, 4), 30, None],
    [date(2015,9, 5), 35, 35],
    [date(2015,9, 6), None, 40]])

# for row in dataframe_to_rows(data, index=False, header=False):
#     ws_data.append(row)

# 컬럼의 너비를 문자에 길이에 맞게 수정후 입력
column_widths= []
for row in dataframe_to_rows(data, index=False, header=False):
    ws_data.append(row)
    for i, cell in enumerate(row):
        cell = str(cell)
        if len(column_widths) > i:
            if len(cell) > column_widths[i]:
                column_widths[i] = len(cell)
        else:
            column_widths += [len(cell)]

for i, column_width in enumerate(column_widths):
    ws_data.column_dimensions[get_column_letter(i+1)].width = column_width



# data header color & border
thin = Side(border_style="thin", color="000000")
for i in range(len(data.columns)):
    data_header = ws_data.cell(row=1,column=i+1)
    data_header.fill = PatternFill("solid", fgColor="DDDDDD")
    data_header.border = Border(top=thin, left=thin, right=thin, bottom=thin)


ws_data.freeze_panes='A2' # A2 위쪽을 고정 
###############################################################

######################### Make Line Chart #########################
line_chart = LineChart()
line_chart.title = "Line Chart Title"
line_chart.style = 12
line_chart.y_axis.title = "Y_title"
line_chart.x_axis.title = "X_title"

chart_data = Reference(ws_data, min_col=2, min_row=2, max_col=3, max_row=8)
line_chart.add_data(chart_data, titles_from_data=True)

# style for chart 
style_fst_val = line_chart.series[0]
style_fst_val.graphicalProperties.line.solidFill = "0000FF"

style_scd_val = line_chart.series[1]
style_scd_val.graphicalProperties.line.solidFill = "FF0000"

ws_report.add_chart(line_chart,"A3")


###############################################################

######################### Column width control #########################

# ws.column_dimensions["A"].width = len(string) 

###############################################################


### sheet 삭제 ###
# wb.remove_sheet()


wb.save(file_nm)
wb.close()

